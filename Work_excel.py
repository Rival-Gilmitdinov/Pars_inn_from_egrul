import openpyxl
from openpyxl import load_workbook
import glob
import os


class Work_excel():
    def __init__(self, file):
        self.file = file
        wb = load_workbook(self.file, read_only=True)
        self.ws = wb.active

    def file_name(self):
        name_split = self.file.filename.rsplit('.')[0]
        return name_split

    def path_dir(self):
        path_dir = f'saving_pdf\\{self.file_name()}'
        return path_dir

    def make_dir(self):
        if self.file_name() not in os.listdir('saving_pdf'):
            os.mkdir(self.path_dir())

    def read_excel(self) -> tuple:
        '''Функия по парсингу таблицы excel
        file : файл, который загружает пользователь
        Return: list_data_inn - list
            список с данными инн
        error_type - dict
            словарь с неверными значениями и текстом ошибки
        error_data - dict
            словарь с неверными значениями и текстом ошибки'''
        list_data_inn = []
        error_data = {}
        # Проходим циклом по значениям таблицы
        for row in self.ws.iter_rows(values_only=True):
            for value in row:
                if type(value) == int or (type(value) == str and value.isdigit()):
                    if len(str(value)) == 10 or len(str(value)) == 12:
                        list_data_inn.append(str(value))
                    else:
                        error_data[value] = 'Неверное количество цифр инн'
                elif value == None:
                    continue
                else:
                    error_data[value] = 'Неверный тип данных'
        return list_data_inn, error_data


class Write_data():
    def __init__(self):
        self.book = openpyxl.Workbook()
        self.sheet = self.book.active

    def write_in_excel(self, data_value, error_data, name_split, error_response) -> None:
        """Функция по записи данных в новую эксель таблицу
        Parameters:
            spisok: list
                Список со словарями, в которых записаны спарсенные данные
            inn: list
                Список с необходимыми инн
            error_type: dict
                Словарь с неверным типом данных:
                    key - данные, которые передал пользователь
                    value - сообщение о неверном данном
            error_data: dict
                Словарь с неверными данными
                    key - данные, которые передал пользователь
                    value - сообщение о неверном данном
        """
        self.sheet['A1'] = 'ИНН'
        self.sheet['B1'] = 'Полное наименование на русском языке'
        self.sheet['C1'] = 'Сведения об уставном капитале / складочном капитале / уставном фонде / паевом фонде'
        self.sheet['D1'] = 'Сведения об основном виде деятельности'
        n = 1
        for value in data_value:
            self.sheet[n][0].value = value['инн']
            self.sheet[n][1].value = value['Полное наименование на русском языке']
            self.sheet[n][2].value = value['Сведения об уставном капитале / складочном капитале / уставном фонде / паевом фонде']
            self.sheet[n][3].value = value['Сведения об основном виде деятельности']
            n += 1
        if error_data:
            for key_error_data, value_error_data in error_data.items():
                self.sheet[n][0].value = key_error_data
                self.sheet[n][1].value = value_error_data
                n += 1
        elif error_response:
            for key, value in error_response.items():
                self.sheet[n][0].value = key
                self.sheet[n][1].value = value
        self.width_column(data_value)
        self.book.save(f'C:\Python\pythonProject\\2025\work_inn\saving_pdf\\{name_split}\\result_data_{name_split}.xlsx')
        self.book.close()

    def width_column(self, data_value):
        max_lenght_inn = 0
        max_lenght_name = 0
        max_lenght_capital = 0
        max_lenght_activity = 0
        for value in data_value:
            if len(str(value['инн'])) >= max_lenght_inn:
                max_lenght_inn = len(str(value['инн']))
            if len(str(value['Полное наименование на русском языке'])) >= max_lenght_name:
                max_lenght_name = len(str(value['Полное наименование на русском языке']))
                print(value['Полное наименование на русском языке'], len(value['Полное наименование на русском языке']))
            if len(str(value['Сведения об уставном капитале / складочном капитале / уставном фонде / паевом фонде'])) >= max_lenght_capital:
                max_lenght_capital = len(str(value['Сведения об уставном капитале / складочном капитале / уставном фонде / паевом фонде']))
            if len(str(value['Сведения об основном виде деятельности'])) >= max_lenght_activity:
                max_lenght_activity = len(str(value['Сведения об основном виде деятельности']))
        self.sheet.column_dimensions['A'].width = max_lenght_inn + 2
        self.sheet.column_dimensions['B'].width = max_lenght_name + 6
        self.sheet.column_dimensions['C'].width = max_lenght_capital + 2
        self.sheet.column_dimensions['D'].width = max_lenght_activity + 2




















