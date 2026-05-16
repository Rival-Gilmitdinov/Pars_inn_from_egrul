from openpyxl import load_workbook
from sqlalchemy import create_engine, Table, func
from sqlalchemy.orm import declarative_base, Session, Mapped
import psycopg2
from config import user, password, port, host, db


# wb = load_workbook('C:\\Users\\79995\OneDrive\Рабочий стол\kkk.xlsx', read_only=True)
# ws = wb.active
#
# for row in ws.iter_rows(values_only=True):
#     for i in range (len(row)):
#         print(row[i])

engine = create_engine(f'postgresql+psycopg2://{user}:{password}@{host}:{port}/{db}')
Base = declarative_base()
list_from_user_file = ["3444051154", "3900005335", "5403336490", "7451053061", "7707009586", "7736222661", "7736617998"]
class Table_data(Base):
    __table__ = Table('data_on_inn', Base.metadata, autoload_with=engine)

with Session(engine) as session:
    results = session.query(Table_data).filter(Table_data.inn_company == func.any(list_from_user_file))
qqq = []
parsing_data = {'инн': None,
                'Полное наименование на русском языке': None,
                'Сведения об уставном капитале / складочном капитале / уставном фонде / паевом фонде': None,
                'Сведения об основном виде деятельности': None}
for i in results:
    parsing_data['инн'] = i.inn_company
    parsing_data['Полное наименование на русском языке'] = i.name
    parsing_data['Сведения об уставном капитале / складочном капитале / уставном фонде / паевом фонде'] = i.capital
    parsing_data['Сведения об основном виде деятельности'] = i.activity
    copy_dict = parsing_data.copy()
    qqq.append(copy_dict)


print(qqq)
#
#
# for i in results:
#     print(i.name)